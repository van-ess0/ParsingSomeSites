import urllib.request
import time
from grab import Grab
import gc
import os
from openpyxl import load_workbook


def filling_into_table(plus=0):
    global number_of_lamps, lamp_type, holder, area, height, length, width, \
        diameter, color_of_base, color_of_plafon, base_material, plafon_material, \
        remote_control, max_power, ws, i
    if number_of_lamps != '':
        ws.cell(column=9+plus, row=i).value = number_of_lamps
    if number_of_lamps != '':
        ws.cell(column=10+plus, row=i).value = style
    if number_of_lamps != '':
        ws.cell(column=11+plus, row=i).value = place_of_usage
    if number_of_lamps != '':
        ws.cell(column=12+plus, row=i).value = lamp_type
    if number_of_lamps != '':
        ws.cell(column=13+plus, row=i).value = holder
    if number_of_lamps != '':
        ws.cell(column=14+plus, row=i).value = area
    if number_of_lamps != '':
        ws.cell(column=15+plus, row=i).value = height
    if number_of_lamps != '':
        ws.cell(column=16+plus, row=i).value = length
    if number_of_lamps != '':
        ws.cell(column=17+plus, row=i).value = width
    if number_of_lamps != '':
        ws.cell(column=18+plus, row=i).value = diameter
    if number_of_lamps != '':
        ws.cell(column=19+plus, row=i).value = color_of_plafon
    if number_of_lamps != '':
        ws.cell(column=20+plus, row=i).value = color_of_base
    if number_of_lamps != '':
        ws.cell(column=21+plus, row=i).value = base_material
    if number_of_lamps != '':
        ws.cell(column=22+plus, row=i).value = plafon_material
    if number_of_lamps != '':
        ws.cell(column=23+plus, row=i).value = remote_control
    if number_of_lamps != '':
        ws.cell(column=24+plus, row=i).value = max_power


def searching_data(data_type, mode=1):
    value = ''
    if mode == 1:   # aspsvet.ru
        for j in range(1, 15):
            try:
                label = g.doc.select('//*[@id="product_addtocart_form"]/div[3]/div/div[7]/table/tbody/tr[{}]/td[1]'.format(j)).text()
                if label == data_type:
                    value = g.doc.select('//*[@id="product_addtocart_form"]/div[3]/div/div[7]/table/tbody/tr[{}]/td[2]'.format(j)).text()
                    break
            except:
                break
    if mode == 2:   # svetodom.ru
        for j in range(1, 15):
            try:
                label = g2.doc.select('//*[@id="description_har"]/table//tr[{}]/td[1]'.format(j)).text()
                if label.startswith(data_type):
                    value = g2.doc.select('//*[@id="description_har"]/table//tr[{}]/td[2]'.format(j)).text()
                    break
            except:
                break
    if mode == 3:   # shop.electra.ru
        for j in range(1, 20):
            try:
                label = g3.doc.select('//table[@class="properties"]//tr[{}]/th'.format(j)).text() + "         "
                if label.startswith(data_type):
                    value = g3.doc.select('//table[@class="properties"]//tr[{}]/td'.format(j)).text()
                    break
            except:
                # break
                print(end='')
    if mode == 4:  # vamsvet.ru
        for j in range(1, 30):
            try:
                label = g4.doc.select('//table[@class="product-tech"]//tr[{}]/th'.format(j)).text()
                if label == data_type:
                    value = g4.doc.select('//table[@class="product-tech"]//tr[{}]/td'.format(j)).text()
                    break
            except:
                break
    # print(data_type + ' ' + value)
    return value


def verify_manufacter(target_manufacter, mode=1):
    value = ''
    if mode == 1:  # aspsvet.ru
        for j in range(1, 15):
            try:
                label = g.doc.select('//*[@id="product_addtocart_form"]/div[3]/div/div[7]/table/tbody/tr[{}]/td[1]'.format(j)).text()
                if label == 'Бренд:':
                    value = g.doc.select('//*[@id="product_addtocart_form"]/div[3]/div/div[7]/table/tbody/tr[{}]/td[2]/a'.format(j)).text().lower()
                    break
            except:
                break

    if mode == 2:   # svetodom.ru
        for j in range(1, 20):
            try:
                label = g2.doc.select('//*[@id="description_har"]/table//tr[{}]/td[1]'.format(j)).text()
                if label.startswith('Производитель'):
                    value = g2.doc.select('//*[@id="description_har"]/table//tr[{}]/td[2]/a'.format(j)).text().lower()
                    break
            except:
                break
    if mode == 3:   # shop.electra.ru
        return True    # TODO: сделать что-то с этим адовым костылём
        try:
            value = g3.doc.select('//span[@class="brand"]').text().lower()
        except:
            print(':(( electra no brand')

    if mode == 4:   #vamsvet.ru
        for j in range(1, 30):
            try:
                label = g4.doc.select('//div[@class="product-info"]/table//tr[{}]/th'.format(j)).text()
                if label == 'Производитель':
                    value = g4.doc.select('//div[@class="product-info"]/table//tr[{}]/td'.format(j)).text().lower()
                    break
            except:
                break
    # print('Target manufacter: ' + target_manufacter.lower())
    # print('Real manufacter: ' + value)
    if target_manufacter.lower() == value:
        # print('OK!!')
        return True
    else:
        return False


def get_from_aspsvet():
    global number_of_lamps, lamp_type, holder, area, height, length, width, \
        diameter, color_of_base, color_of_plafon, base_material, plafon_material, \
        remote_control, max_power, g, article, manufacter
    try:
        g.doc.set_input_by_id('search', str(article.value))
        g.doc.submit()

        g.go(g.doc.select('//ul/li[1]/div/div/a/@href').text())
    except:
        print(':( aspsvet')
        return
    # Verify manufacter
    if verify_manufacter(manufacter):
        # Number of lamps
        number_of_lamps = searching_data('Кол-во лампочек:')

        # Style
        # TODO: Parse Styles anywhere Oo

        # Place of usage
        # TODO: Parse Place of usage anywhere Oo

        # Lamp type
        lamp_type = searching_data('Тип лампочек:')

        # Holder
        holder = searching_data('Цоколь:')

        # Area
        area = ''.join(c for c in searching_data('Площадь освещения:') if c.isdigit())

        # Height
        height = ''.join(c for c in searching_data('Высота:') if c.isdigit())

        # Length
        length = ''.join(c for c in searching_data('Длина:') if c.isdigit())

        # Width
        width = ''.join(c for c in searching_data('Ширина:') if c.isdigit())

        # Diameter
        diameter = ''.join(c for c in searching_data('Диаметр:') if c.isdigit())

        # Color of plafon
        color_of_plafon = searching_data('Цвет плафона:')

        # Color of base
        color_of_base = searching_data('Цвет арматуры:')

        # Base material

        # Plafon material

        # Remote control

        # Max power of lamp
        max_power = ''.join(c for c in searching_data('Мощность:') if c.isdigit())

    g.go("http://aspsvet.ru/")


def get_from_svetodom():
    global number_of_lamps, lamp_type, holder, area, height, length, width,\
            diameter, color_of_base, color_of_plafon, base_material, plafon_material,\
            remote_control, max_power, g2, article, manufacter
    try:
        g2.go('http://www.svetodom.ru/search/?ukey=search&searchstring={}'.format(str(article.value)))
    # //*[@id="tbl_cat_prd"]/tbody/tr/td/div/div[1]/a
        g2.go('http://www.svetodom.ru' + g2.doc.select('//*[@id="tbl_cat_prd"]//tr/td/div/div/a/@href').text())
    except:
         print(':( svetodom')
         return
        # Do nothing
    # Verify manufacter
    if verify_manufacter(manufacter, mode=2):
        # Number of lamps
        if number_of_lamps == '':
            number_of_lamps = searching_data('Количество ламп', mode=2)

        # Style
        # if style == '':
        #     # TODO: Parse Styles anywhere Oo
        #
        # # Place of usage
        # if place_of_usage == '':
        #     # TODO: Parse Place of usage anywhere Oo

        # Lamp type
        if lamp_type == '':
            lamp_type = searching_data('Тип лампы', mode=2)

        # Holder
        if holder == '':
            holder = searching_data('Тип цоколя', mode=2)

        # Area
        if area == '':
            area = searching_data('S освещ. до', mode=2)

        # Height
        if height == '':
            height = searching_data('Высота', mode=2)

        # Length
        if length == '':
            length = searching_data('Длина', mode=2)

        # Width
        if width == '':
            width = searching_data('Ширина', mode=2)

        # Diameter
        if diameter == '':
            diameter = searching_data('Диаметр', mode=2)

        # Color of plafon
        if color_of_plafon == '':
            color_of_plafon = searching_data('Плафон', mode=2)

        # Color of base
        if color_of_base == '':
            color_of_base = searching_data('Цвет арматуры', mode=2)

        # Base material
        if base_material == '':
            base_material = searching_data('Материалы', mode=2)

        # Plafon material
        # if plafon_material == '':
        #     plafon_material = searching_data('Материалы', mode=2)
        try:
            base_material, plafon_material = base_material.split('/', maxsplit=1)
        except ValueError:
            try:
                base_material, plafon_material = base_material.split(', ', maxsplit=1)
            except ValueError:
                plafon_material = ''

        # Remote control
        # if remote_control == '':
        #     remote_control = searching_data('', mode=2)

        # Max power of lamp
        if max_power == '':
            max_power = searching_data('MAX мощность ламп', mode=2)

    g2.go("http://www.svetodom.ru/")


def get_from_electra():
    global number_of_lamps, lamp_type, holder, area, height, length, width, \
        diameter, color_of_base, color_of_plafon, base_material, plafon_material, \
        remote_control, max_power, g3, article, manufacter, model
    try:
        g3.doc.set_input_by_xpath('//div[2]/div[1]/div[5]/form/input[1]', model)
        g3.doc.submit()
        # //div[2]/div[4]/div/div[2]/div/a[2]/@href
        g3.go('http://shop.electra.ru' + g3.doc.select('//div[@class="search-item nofloat"]/a/@href').text())
    except:
        print(':( electra')
        return
    # Verify manufacter
    if verify_manufacter(manufacter, mode=3):

        # if manufacter == 'Electra':
        #     ws.cell(column=3, row=i).value = g3.doc.select('//span[@class="brand"]').text()
        # Number of lamps
        if number_of_lamps == '':
            number_of_lamps = searching_data('Количество ламп', mode=3)

        # Style
        # if style == '':
        #     # TODO: Parse Styles anywhere Oo
        #
        # # Place of usage
        # if place_of_usage == '':
        #     # TODO: Parse Place of usage anywhere Oo

        # Lamp type
        if lamp_type == '':
            lamp_type = searching_data('Тип лампы', mode=3)

        # Holder
        if holder == '':
            holder = searching_data('Цоколь', mode=3)

        # Area
        if area == '':
            area = searching_data('S освещ. до', mode=3)

        # Height
        if height == '':
            height = searching_data('Высота', mode=3)

        # Length
        if length == '':
            length = searching_data('Длина', mode=3)

        # Width
        if width == '':
            width = searching_data('Ширина', mode=3)

        # Diameter
        if diameter == '':
            diameter = searching_data('Диаметр', mode=3)

        # Color of plafon
        if color_of_plafon == '':
            color_of_plafon = searching_data('Цвет плафона', mode=3)

        # Color of base
        if color_of_base == '':
            color_of_base = searching_data('Цвет основания', mode=3)

        # Base material
        if base_material == '':
            base_material = searching_data('Материал основания', mode=3)

        # Plafon material
        if plafon_material == '':
            plafon_material = searching_data('Материал плафона', mode=3)

        # Remote control
        if remote_control == '':
            remote_control = searching_data('Наличие ПДУ', mode=3)

        # Max power of lamp
        if max_power == '':
            max_power = searching_data('Потребляемая мощность', mode=3)

    g3.go('http://shop.electra.ru/')


def get_from_vamsvet():
    global number_of_lamps, lamp_type, holder, area, height, length, width, \
        diameter, color_of_base, color_of_plafon, base_material, plafon_material, \
        remote_control, max_power, g4, article, manufacter, place_of_usage, style
    try:
        g4.doc.set_input_by_xpath('//div[@class="header-menu-search"]/form/input', str(article.value))
        g4.doc.submit()
        g4.go('http://www.vamsvet.ru' + g4.doc.select('//div[@class="prods-list"]/div[3]/div/a/@href').text())
    except:
        print(':( vamsvet')
        return
    # Verify manufacter
    if verify_manufacter(manufacter, mode=4):
        # Number of lamps
        if number_of_lamps == '':
            number_of_lamps = searching_data('Количество ламп', mode=4)

        # Style
        if style == '':
            style = searching_data('Стиль', mode=4)

        # Place of usage
        if place_of_usage == '':
            place_of_usage = searching_data('Интерьер', mode=4)

        # Lamp type
        if lamp_type == '':
            lamp_type = searching_data('Тип лампочки (основной)', mode=4)

        # Holder
        if holder == '':
            holder = searching_data('Тип цоколя', mode=4)

        # Area
        if area == '':
            area = searching_data('Площадь освещения, м2', mode=4) + 'кв м'

        # Height
        if height == '':
            height = searching_data('Высота, мм', mode=4) + 'мм'

        # Length
        if length == '':
            length = searching_data('Длина, мм', mode=4) + 'мм'

        # Width
        if width == '':
            width = searching_data('Ширина, мм', mode=4) + 'мм'

        # Diameter
        if diameter == '':
            diameter = searching_data('Диаметр, мм', mode=4) + 'мм'

        # Color of plafon
        if color_of_plafon == '':
            color_of_plafon = searching_data('Цвет плафонов', mode=4)

        # Color of base
        if color_of_base == '':
            color_of_base = searching_data('Цвет арматуры', mode=4)

        # Base material
        if base_material == '':
            base_material = searching_data('Материал арматуры', mode=4)

        # Plafon material
        if plafon_material == '':
            plafon_material = searching_data('Материал плафонов', mode=4)

        # Remote control
        # if remote_control == '':
        #     remote_control = searching_data('', mode=4)

        # Max power of lamp
        if max_power == '':
            max_power = searching_data('Общая мощность, W', mode=4)

    g4.go('http://www.vamsvet.ru/')

gc.enable()

g = Grab()
g.go('http://aspsvet.ru/')

g2 = Grab()
g2.go('http://www.svetodom.ru/')

g3 = Grab()
g3.setup(cookies={'BITRIX_SM_NX_OFFICE_id': '80438'})
g3.go('http://shop.electra.ru/')

g4 = Grab()
g4.go('http://www.vamsvet.ru/')

g5 = Grab()
g5.go('http://magia-sveta.ru/')

wb = load_workbook('Edited Sortirovka.xlsx')

# ws = wb.get_sheet_by_name('Доп.освещение Торшеры')
# for i in range(2, 156):   # 37
#     article = ws.cell(column=2, row=i)
#     manufacter = ws.cell(column=4, row=i).value
#     model = ws.cell(column=3, row=i).value
#
#     # Initialising variables
#     number_of_lamps = ''
#     style = ''
#     place_of_usage = ''
#     lamp_type = ''
#     holder = ''
#     area = ''
#     height = ''
#     length = ''
#     width = ''
#     diameter = ''
#     color_of_plafon = ''
#     color_of_base = ''
#     base_material = ''
#     plafon_material = ''
#     remote_control = ''
#     max_power = ''
#
#     print('Number: ' + str(i))
#     print('Article: ' + str(article.value))
#
#     get_from_aspsvet()
#     get_from_svetodom()
#     get_from_electra()
#     get_from_vamsvet()
#
#     filling_into_table()
#
#     gc.collect()
# wb.save('Edited Sortirovka.xlsx')
#
# ws = wb.get_sheet_by_name('Бра')
# for i in range(207, 842):   # 37
#     article = ws.cell(column=2, row=i)
#     manufacter = ws.cell(column=4, row=i).value
#     model = ws.cell(column=3, row=i).value
#
#     # Initialising variables
#     number_of_lamps = ''
#     style = ''
#     place_of_usage = ''
#     lamp_type = ''
#     holder = ''
#     area = ''
#     height = ''
#     length = ''
#     width = ''
#     diameter = ''
#     color_of_plafon = ''
#     color_of_base = ''
#     base_material = ''
#     plafon_material = ''
#     remote_control = ''
#     max_power = ''
#
#     print('Number: ' + str(i))
#     print('Article: ' + str(article.value))
#
#     get_from_aspsvet()
#     get_from_svetodom()
#     get_from_electra()
#     get_from_vamsvet()
#
#     filling_into_table()
#
#     gc.collect()
# wb.save('Edited Sortirovka.xlsx')
#
# ws = wb.get_sheet_by_name('Люстры')
# for i in range(2, 2102):   # 37
#     article = ws.cell(column=2, row=i)
#     manufacter = ws.cell(column=8, row=i).value
#     model = ws.cell(column=3, row=i).value
#
#     # Initialising variables
#     number_of_lamps = ''
#     style = ''
#     place_of_usage = ''
#     lamp_type = ''
#     holder = ''
#     area = ''
#     height = ''
#     length = ''
#     width = ''
#     diameter = ''
#     color_of_plafon = ''
#     color_of_base = ''
#     base_material = ''
#     plafon_material = ''
#     remote_control = ''
#     max_power = ''
#
#     print('Number: ' + str(i))
#     print('Article: ' + str(article.value))
#
#     get_from_aspsvet()
#     get_from_svetodom()
#     get_from_electra()
#     get_from_vamsvet()
#
#     filling_into_table()
#
#     gc.collect()
# wb.save('Edited Sortirovka.xlsx')
#
# ws = wb.get_sheet_by_name('Светильники')
# for i in range(2, 1589):   # 1589
#     article = ws.cell(column=2, row=i)
#     manufacter = ws.cell(column=4, row=i).value
#     model = ws.cell(column=3, row=i).value
#
#     # Initialising variables
#     number_of_lamps = ''
#     style = ''
#     place_of_usage = ''
#     lamp_type = ''
#     holder = ''
#     area = ''
#     height = ''
#     length = ''
#     width = ''
#     diameter = ''
#     color_of_plafon = ''
#     color_of_base = ''
#     base_material = ''
#     plafon_material = ''
#     remote_control = ''
#     max_power = ''
#
#     print('Number: ' + str(i))
#     print('Article: ' + str(article.value))
#
#     get_from_aspsvet()
#     get_from_svetodom()
#     get_from_electra()
#     get_from_vamsvet()
#
#     filling_into_table(plus=1)
#
#     gc.collect()
# wb.save('Edited Sortirovka.xlsx')

# ws = wb.get_sheet_by_name('Доп.освещение')
# for i in range(324, 384):   # 456
#     article = ws.cell(column=2, row=i)
#     manufacter = ws.cell(column=4, row=i).value
#     model = ws.cell(column=3, row=i).value
#
#     # Initialising variables
#     number_of_lamps = ''
#     style = ''
#     place_of_usage = ''
#     lamp_type = ''
#     holder = ''
#     area = ''
#     height = ''
#     length = ''
#     width = ''
#     diameter = ''
#     color_of_plafon = ''
#     color_of_base = ''
#     base_material = ''
#     plafon_material = ''
#     remote_control = ''
#     max_power = ''
#
#     print('Number: ' + str(i))
#     print('Article: ' + str(article.value))
#
#     get_from_aspsvet()
#     get_from_svetodom()
#     get_from_electra()
#     get_from_vamsvet()
#
#     filling_into_table(plus=1)
#
#     gc.collect()
# wb.save('Edited Sortirovka.xlsx')
